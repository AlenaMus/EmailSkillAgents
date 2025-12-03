"""
Excel Export Module

This module handles Excel file generation with smart filename generation
based on active filters and organized folder structure.
"""

import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export emails to Excel with smart naming and formatting"""

    DEFAULT_EXPORTS_DIR = "./exports"

    def __init__(self, exports_dir: Optional[str] = None):
        """
        Initialize Excel exporter

        Args:
            exports_dir: Directory to save Excel files (default: ./exports)
        """
        self.exports_dir = Path(exports_dir) if exports_dir else Path(self.DEFAULT_EXPORTS_DIR)

    def ensure_exports_folder(self):
        """Create exports folder if it doesn't exist"""
        self.exports_dir.mkdir(parents=True, exist_ok=True)

    def generate_filename(self, filters: Dict[str, str]) -> str:
        """
        Generate smart filename based on active filters

        Args:
            filters: Dictionary of active filters

        Returns:
            Generated filename in format: DD.MM.YYYY_HHMMSS_mails_<filter_description>.xlsx

        Examples:
            - 22.11.2024_143022_mails_by_tag_Important.xlsx
            - 22.11.2024_143022_mails_from_folder_Work.xlsx
            - 22.11.2024_143530_mails_by_tag_Work_from_sender_john@example.com.xlsx
        """
        # Get current date and time
        now = datetime.now()
        date_str = now.strftime("%d.%m.%Y")  # DD.MM.YYYY
        time_str = now.strftime("%H%M%S")    # HHMMSS

        # Build filter description
        filter_parts = []

        if filters.get('label'):
            label = self._sanitize_filename(filters['label'])
            filter_parts.append(f"by_label_{label}")

        if filters.get('folder'):
            folder = self._sanitize_filename(filters['folder'])
            filter_parts.append(f"from_folder_{folder}")

        if filters.get('tag'):
            tag = self._sanitize_filename(filters['tag'])
            filter_parts.append(f"by_tag_{tag}")

        if filters.get('from'):
            sender = self._sanitize_filename(filters['from'])
            filter_parts.append(f"from_sender_{sender}")

        if filters.get('to'):
            recipient = self._sanitize_filename(filters['to'])
            filter_parts.append(f"to_recipient_{recipient}")

        if filters.get('subject'):
            subject = self._sanitize_filename(filters['subject'])
            filter_parts.append(f"by_subject_{subject}")

        # Combine all parts
        if not filter_parts:
            filter_parts = ["all_mails"]

        filter_desc = "_".join(filter_parts)

        # Limit total filename length (reserve space for date, time, and extension)
        max_filter_length = 150
        if len(filter_desc) > max_filter_length:
            filter_desc = filter_desc[:max_filter_length]

        # Generate filename
        filename = f"{date_str}_{time_str}_mails_{filter_desc}.xlsx"

        return filename

    def _sanitize_filename(self, text: str, max_length: int = 50) -> str:
        """
        Remove special characters and limit length for filename

        Args:
            text: Text to sanitize
            max_length: Maximum length for the sanitized text

        Returns:
            Sanitized text safe for filenames
        """
        # Remove or replace special characters
        text = re.sub(r'[<>:"/\\|?*@]', '_', text)
        # Replace spaces with underscores
        text = text.replace(' ', '_')
        # Remove multiple underscores
        text = re.sub(r'_+', '_', text)
        # Limit length
        if len(text) > max_length:
            text = text[:max_length]
        # Remove trailing underscores
        text = text.strip('_')
        return text

    def create_excel(
        self,
        emails: List[Dict],
        filters: Dict[str, str],
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate Excel file from email data

        Args:
            emails: List of email dictionaries
            filters: Active filters for filename generation
            output_path: Custom output path (overrides auto-generation)

        Returns:
            Path to the created Excel file
        """
        # Ensure exports folder exists
        self.ensure_exports_folder()

        # Determine output path
        if output_path:
            output_file = Path(output_path)
        else:
            filename = self.generate_filename(filters)
            output_file = self.exports_dir / filename

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Emails"

        # Header row
        headers = ["ID", "Date", "Subject", "URL", "Status"]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sort emails by date (newest first)
        sorted_emails = sorted(emails, key=lambda x: x.get('date', datetime.now()), reverse=True)

        # Data rows
        for email in sorted_emails:
            # Generate unique UUID
            email_id = str(uuid.uuid4())

            # Format date
            date = email.get('date')
            if date:
                date_str = date.strftime("%Y-%m-%d %H:%M:%S")
            else:
                date_str = ""

            # Get subject
            subject = email.get('subject', '(No Subject)')

            # Format URLs (comma-separated)
            urls = email.get('urls', [])
            url_str = ', '.join(urls) if urls else ""

            # Add row with Status = "ready"
            row = [email_id, date_str, subject, url_str, "ready"]
            ws.append(row)

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save workbook
        wb.save(output_file)

        return str(output_file)

    def _adjust_column_widths(self, ws):
        """
        Auto-adjust column widths based on content

        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                column_letter = cell.column_letter
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass

            # Set adjusted width (with limits)
            adjusted_width = min(max_length + 2, 100)  # Max 100 characters wide
            adjusted_width = max(adjusted_width, 10)   # Min 10 characters wide

            if column_letter:
                ws.column_dimensions[column_letter].width = adjusted_width


def export_emails_to_excel(
    emails: List[Dict],
    filters: Dict[str, str],
    output_path: Optional[str] = None,
    exports_dir: Optional[str] = None
) -> str:
    """
    Helper function to export emails to Excel

    Args:
        emails: List of email dictionaries
        filters: Active filters for filename generation
        output_path: Custom output path (optional)
        exports_dir: Custom exports directory (optional)

    Returns:
        Path to the created Excel file
    """
    exporter = ExcelExporter(exports_dir)
    return exporter.create_excel(emails, filters, output_path)
