"""
Excel Manager - Handles reading and writing Excel files with greetings
"""

import openpyxl
from openpyxl.styles import Alignment
from pathlib import Path
from typing import List, Dict, Any, Optional
from .config import EXCEL_COLUMNS, NEW_COLUMNS


class ExcelManager:
    """
    Manages Excel I/O operations for greetings agent
    """

    def read_graded_excel(self, file_path: str) -> tuple[openpyxl.Workbook, List[Dict[str, Any]]]:
        """
        Read graded Excel from Repository Analyzer

        Args:
            file_path: Path to graded Excel file

        Returns:
            Tuple of (workbook, list of student records)

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If Excel structure is invalid
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Load workbook
        wb = openpyxl.load_workbook(file_path)

        # Get "Graded Results" sheet
        if "Graded Results" not in wb.sheetnames:
            raise ValueError("Excel file missing 'Graded Results' sheet")

        ws = wb["Graded Results"]

        # Read header row
        headers = [cell.value for cell in ws[1]]

        # Validate required columns
        if 'Grade' not in headers:
            raise ValueError("Excel file missing 'Grade' column")

        # Check if agent coordination Status column exists and validate all rows are "ready"
        # Find the RIGHTMOST Status column (agent coordination column)
        if 'Status' in headers:
            # Get all Status column indices
            status_indices = [i for i, h in enumerate(headers) if h == 'Status']
            if status_indices:
                # Use the rightmost Status column (agent coordination)
                status_col_idx = status_indices[-1] + 1
                all_ready = True
                not_ready_count = 0
                for row_idx in range(2, ws.max_row + 1):
                    status_value = ws.cell(row=row_idx, column=status_col_idx).value
                    if status_value != 'ready':
                        all_ready = False
                        not_ready_count += 1

                if not all_ready:
                    raise ValueError(
                        f"Cannot process: {not_ready_count} row(s) have Status != 'ready'. "
                        f"Previous agent (Repository Analyzer) must complete successfully before this agent can run."
                    )

        # Read all student records
        students = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value

            # Add row index for later reference
            row_data['_row_index'] = row_idx
            students.append(row_data)

        return wb, students

    def write_with_greetings(self,
                           original_wb: openpyxl.Workbook,
                           students: List[Dict[str, Any]],
                           greetings: List[tuple[str, Optional[str]]],
                           output_path: str):
        """
        Write Excel with greetings columns added

        Args:
            original_wb: Original workbook from read_graded_excel
            students: List of student records
            greetings: List of (greeting_text, persona_name) tuples
            output_path: Path to output Excel file
        """
        # Get "Graded Results" sheet
        ws = original_wb["Graded Results"]

        # Get current headers
        headers = [cell.value for cell in ws[1]]

        # Add new column headers
        next_col = len(headers) + 1
        ws.cell(row=1, column=next_col, value="Personalized Greeting")
        ws.cell(row=1, column=next_col + 1, value="Greeting Persona")
        ws.cell(row=1, column=next_col + 2, value="Status")  # Agent coordination Status

        # Add greetings to each row
        for idx, (student, (greeting, persona)) in enumerate(zip(students, greetings)):
            row_idx = student['_row_index']

            # Add greeting
            greeting_cell = ws.cell(row=row_idx, column=next_col, value=greeting)
            greeting_cell.alignment = Alignment(
                wrap_text=True,
                vertical='top',
                horizontal='left'
            )

            # Add persona name
            ws.cell(row=row_idx, column=next_col + 1, value=persona if persona else "N/A")

            # Add Status = "ready" after processing
            ws.cell(row=row_idx, column=next_col + 2, value="ready")

        # Set column widths
        ws.column_dimensions[openpyxl.utils.get_column_letter(next_col)].width = 60
        ws.column_dimensions[openpyxl.utils.get_column_letter(next_col + 1)].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(next_col + 2)].width = 15

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        original_wb.save(output_path)

    def validate_input_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Validate input Excel structure before processing

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary with validation results
        """
        result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'student_count': 0,
            'graded_count': 0,
            'error_count': 0
        }

        try:
            wb, students = self.read_graded_excel(file_path)

            result['student_count'] = len(students)

            # Count graded vs error rows
            for student in students:
                grade = student.get('Grade')
                if grade is not None and grade != '':
                    result['graded_count'] += 1
                else:
                    result['error_count'] += 1

            if result['graded_count'] == 0:
                result['valid'] = False
                result['errors'].append("No students with valid grades found")

        except FileNotFoundError as e:
            result['valid'] = False
            result['errors'].append(str(e))
        except ValueError as e:
            result['valid'] = False
            result['errors'].append(str(e))
        except Exception as e:
            result['valid'] = False
            result['errors'].append(f"Unexpected error: {str(e)}")

        return result
