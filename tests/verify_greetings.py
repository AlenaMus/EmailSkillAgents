import openpyxl

# Load the output file
wb = openpyxl.load_workbook(r'C:\AIDevelopmentCourse\L-19\EmailSkillAgents\greetings_results\homework_emails_with_greetings.xlsx')
ws = wb.active

print('Sheet names:', wb.sheetnames)
print('\nColumns:', [cell.value for cell in ws[1]])
print('\nTotal rows:', ws.max_row)

print('\n' + '='*80)
print('SAMPLE DATA (First 2 students)')
print('='*80)

headers = [cell.value for cell in ws[1]]

for row_idx in range(2, min(4, ws.max_row+1)):
    print(f'\nStudent {row_idx-1}:')
    for col_idx, header in enumerate(headers, start=1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if header in ['Personalized Greeting', 'Greeting Persona']:
            print(f'  {header}: {value}')
        elif header == 'Grade':
            print(f'  {header}: {value}')
