import openpyxl

wb = openpyxl.load_workbook('output/homework_emails_graded.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print('Headers:', headers)

# Find all Status columns
status_indices = [i for i, h in enumerate(headers, 1) if h == 'Status']
print(f'Status column indices: {status_indices}')

for status_col_idx in status_indices:
    print(f'\nStatus column at index {status_col_idx}:')
    for i in range(1, min(8, ws.max_row+1)):
        value = ws.cell(row=i, column=status_col_idx).value
        print(f'  Row {i}: {value!r}')
