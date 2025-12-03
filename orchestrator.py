"""
Email Skill Agents Orchestrator
================================

Main orchestrator for running the 4-agent email grading workflow.

Agent Workflow:
1. GmailAgent - Export homework emails to Excel
2. Repository Analyzer - Analyze repos and calculate grades
3. Personalized Greetings Agent - Add motivational feedback
4. Evaluation Gmail Sender - Send feedback to students

Author: AI Development Course - Lesson 19
"""

import os
import sys
import shutil
from pathlib import Path
from typing import Optional, List
import subprocess
import openpyxl


class AgentOrchestrator:
    """Orchestrator for managing the 4-agent email grading workflow."""

    # Paths
    PROJECT_ROOT = Path(__file__).parent
    EXPORTS_DIR = PROJECT_ROOT / "exports"
    OUTPUT_DIR = PROJECT_ROOT / "output"
    GREETINGS_RESULTS_DIR = PROJECT_ROOT / "greetings_results"
    EMAIL_DRAFTS_DIR = PROJECT_ROOT / "email_drafts"

    # Expected file paths
    GMAIL_OUTPUT = EXPORTS_DIR / "homework_emails.xlsx"
    REPO_ANALYZER_OUTPUT = OUTPUT_DIR / "homework_emails_graded.xlsx"
    GREETINGS_OUTPUT = GREETINGS_RESULTS_DIR / "homework_emails_with_greetings.xlsx"

    # Default email for testing
    DEFAULT_EMAIL = "alona.musiyko@gmail.com"

    def __init__(self):
        """Initialize orchestrator."""
        self.ensure_directories()

    def ensure_directories(self):
        """Ensure all required directories exist."""
        self.EXPORTS_DIR.mkdir(exist_ok=True)
        self.OUTPUT_DIR.mkdir(exist_ok=True)
        self.GREETINGS_RESULTS_DIR.mkdir(exist_ok=True)
        self.EMAIL_DRAFTS_DIR.mkdir(exist_ok=True)

    def print_header(self, text: str):
        """Print formatted header."""
        print("\n" + "=" * 70)
        print(f"  {text}")
        print("=" * 70)

    def print_status(self, status: str, message: str):
        """Print status message."""
        symbols = {
            "success": "[OK]",
            "error": "[ERROR]",
            "info": "[INFO]",
            "warning": "[WARN]"
        }
        symbol = symbols.get(status, "[*]")
        print(f"{symbol} {message}")

    def check_file_exists(self, file_path: Path) -> bool:
        """Check if file exists."""
        return file_path.exists() and file_path.is_file()

    def check_status_ready(self, excel_path: Path, column_name: str = "Status") -> tuple[bool, str]:
        """
        Check if Excel file has Status = "ready" in all rows.

        Args:
            excel_path: Path to Excel file
            column_name: Name of status column to check

        Returns:
            Tuple of (all_ready: bool, message: str)
        """
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Find Status column (rightmost if multiple)
            status_indices = [i for i, h in enumerate(headers) if h == column_name]
            if not status_indices:
                return False, f"No '{column_name}' column found in Excel file"

            status_col_idx = status_indices[-1]  # Get rightmost Status column

            # Check all rows
            not_ready_count = 0
            total_rows = ws.max_row - 1  # Exclude header

            for row in ws.iter_rows(min_row=2, values_only=True):
                if status_col_idx < len(row):
                    status_value = row[status_col_idx]
                    if status_value != "ready":
                        not_ready_count += 1

            if not_ready_count > 0:
                return False, f"{not_ready_count}/{total_rows} rows not ready (Status != 'ready')"

            wb.close()
            return True, f"All {total_rows} rows ready"

        except Exception as e:
            return False, f"Error checking status: {str(e)}"

    def run_agent_1_gmail(
        self,
        label: Optional[str] = None,
        subject: Optional[str] = None,
        after: Optional[str] = None,
        before: Optional[str] = None,
        newer_than: Optional[str] = None
    ) -> bool:
        """
        Run Agent 1: GmailAgent - Export homework emails to Excel.

        Args:
            label: Gmail label to filter by
            subject: Subject text to filter by
            after: Filter emails after this date (YYYY-MM-DD)
            before: Filter emails before this date (YYYY-MM-DD)
            newer_than: Filter emails newer than (e.g., 7d, 2m, 1y)

        Returns:
            True if successful, False otherwise
        """
        self.print_header("Agent 1: GmailAgent - Export Homework Emails")

        # If no filters provided, ask user
        if not any([label, subject, after, before, newer_than]):
            print("\nPlease specify how to filter emails:")
            print("You can use any combination of filters below")
            print("\nFilter Options:")
            print("  Label:    'homework', 'lesson19', 'assignments'")
            print("  Subject:  'Homework 19', 'Assignment', 'Project'")
            print("  After:    '2025-12-01' (emails after this date)")
            print("  Before:   '2025-12-03' (emails before this date)")
            print("  Newer:    '7d' (last 7 days), '2m' (last 2 months)")
            print()

            label = input("Enter Gmail label (or press Enter to skip): ").strip()
            subject = input("Enter subject text (or press Enter to skip): ").strip()
            after = input("Enter after date YYYY-MM-DD (or press Enter to skip): ").strip()
            before = input("Enter before date YYYY-MM-DD (or press Enter to skip): ").strip()
            newer_than = input("Enter newer than (e.g., 7d, 2m) (or press Enter to skip): ").strip()

            if not any([label, subject, after, before, newer_than]):
                self.print_status("warning", "No filters specified. Will export all emails.")
                proceed = input("Continue? (yes/no): ").strip().lower()
                if proceed not in ["yes", "y"]:
                    self.print_status("info", "Cancelled by user")
                    return False

        try:
            # Build gmailagent command (run as Python module)
            cmd = [sys.executable, "-m", "gmailagent", "export"]

            if label:
                cmd.extend(["--label", label])
                self.print_status("info", f"Filter by label: {label}")

            if subject:
                cmd.extend(["--subject", subject])
                self.print_status("info", f"Filter by subject: {subject}")

            if after:
                cmd.extend(["--after", after])
                self.print_status("info", f"Filter emails after: {after}")

            if before:
                cmd.extend(["--before", before])
                self.print_status("info", f"Filter emails before: {before}")

            if newer_than:
                cmd.extend(["--newer-than", newer_than])
                self.print_status("info", f"Filter emails newer than: {newer_than}")

            self.print_status("info", "Running GmailAgent...")
            print(f"\nCommand: {' '.join(cmd)}")
            print("This may take a few moments...\n")

            # Run gmailagent
            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                # Check if no emails were found
                if "No emails found" in result.stdout or "No messages found" in result.stdout:
                    self.print_status("warning", "GmailAgent completed but found no emails")
                    self.print_status("info", "No emails match your filters. Try different label/subject.")
                    print(f"\nOutput:\n{result.stdout}")
                    return False

                self.print_status("success", "GmailAgent export completed")

                # Check if output exists
                # Look for most recent file in exports directory
                if self.EXPORTS_DIR.exists():
                    excel_files = list(self.EXPORTS_DIR.glob("*.xlsx"))
                    if excel_files:
                        # Get most recent file
                        latest_file = max(excel_files, key=lambda p: p.stat().st_mtime)
                        self.print_status("success", f"Found Excel file: {latest_file.name}")

                        # Copy to standard location for next agent
                        shutil.copy(latest_file, self.GMAIL_OUTPUT)
                        self.print_status("success", f"Copied to: {self.GMAIL_OUTPUT}")

                        # Check status
                        ready, msg = self.check_status_ready(self.GMAIL_OUTPUT)
                        if ready:
                            self.print_status("success", f"Status check: {msg}")
                            return True
                        else:
                            self.print_status("warning", f"Status check: {msg}")
                            return True  # Still return True as file exists
                    else:
                        self.print_status("error", "No Excel files found in exports directory")
                        self.print_status("info", "The export may have completed but found no emails")
                        if result.stdout:
                            print(f"\nOutput:\n{result.stdout}")
                        return False
                else:
                    self.print_status("error", "Exports directory not found")
                    return False
            else:
                self.print_status("error", "GmailAgent export failed")
                if result.stdout:
                    print(f"\nOutput:\n{result.stdout}")
                if result.stderr:
                    print(f"\nError details:\n{result.stderr}")
                return False

        except ModuleNotFoundError:
            self.print_status("error", "gmailagent module not found. Is it installed?")
            self.print_status("info", "Try: pip install -e .")
            return False
        except FileNotFoundError:
            self.print_status("error", "Python executable not found")
            return False
        except Exception as e:
            self.print_status("error", f"Failed to run GmailAgent: {str(e)}")
            return False

    def run_agent_2_repo_analyzer(self) -> bool:
        """
        Run Agent 2: Repository Analyzer - Analyze repos and calculate grades.

        Returns:
            True if successful, False otherwise
        """
        self.print_header("Agent 2: Repository Analyzer - Analyze Repositories")

        # Check if input file exists
        if not self.check_file_exists(self.GMAIL_OUTPUT):
            self.print_status("error", f"Input file not found: {self.GMAIL_OUTPUT}")
            self.print_status("info", "Please run Agent 1 first")
            return False

        # Check if input is ready
        ready, msg = self.check_status_ready(self.GMAIL_OUTPUT)
        if not ready:
            self.print_status("error", f"Input not ready: {msg}")
            return False

        self.print_status("success", f"Input ready: {msg}")

        try:
            # Run repository analyzer
            self.print_status("info", "Running Repository Analyzer...")

            cmd = [
                sys.executable,
                "-m", "repo_analyzer.cli",
                "analyze",
                "--input", str(self.GMAIL_OUTPUT),
                "--output", str(self.REPO_ANALYZER_OUTPUT)
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                self.print_status("success", "Repository analysis completed")

                # Check output file
                if self.check_file_exists(self.REPO_ANALYZER_OUTPUT):
                    # Check status
                    ready, msg = self.check_status_ready(self.REPO_ANALYZER_OUTPUT)
                    if ready:
                        self.print_status("success", f"Output status: {msg}")
                    else:
                        self.print_status("warning", f"Output status: {msg}")

                    return True
                else:
                    self.print_status("error", "Output file not created")
                    return False
            else:
                self.print_status("error", "Repository analysis failed")
                if result.stderr:
                    print(f"\nError details:\n{result.stderr}")
                return False

        except Exception as e:
            self.print_status("error", f"Failed to run Repository Analyzer: {str(e)}")
            return False

    def run_agent_3_greetings(self) -> bool:
        """
        Run Agent 3: Personalized Greetings Agent - Add motivational feedback.

        Returns:
            True if successful, False otherwise
        """
        self.print_header("Agent 3: Personalized Greetings Agent - Add Feedback")

        # Check if input file exists
        if not self.check_file_exists(self.REPO_ANALYZER_OUTPUT):
            self.print_status("error", f"Input file not found: {self.REPO_ANALYZER_OUTPUT}")
            self.print_status("info", "Please run Agent 2 first")
            return False

        # Check if input is ready
        ready, msg = self.check_status_ready(self.REPO_ANALYZER_OUTPUT)
        if not ready:
            self.print_status("error", f"Input not ready: {msg}")
            return False

        self.print_status("success", f"Input ready: {msg}")

        try:
            # Run greetings agent
            self.print_status("info", "Running Personalized Greetings Agent...")

            cmd = [
                sys.executable,
                "-m", "greetings_grades_agent.cli",
                "greet",
                "--input", str(self.REPO_ANALYZER_OUTPUT),
                "--output", str(self.GREETINGS_OUTPUT)
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                self.print_status("success", "Greetings generation completed")

                # Check output file
                if self.check_file_exists(self.GREETINGS_OUTPUT):
                    # Check status
                    ready, msg = self.check_status_ready(self.GREETINGS_OUTPUT)
                    if ready:
                        self.print_status("success", f"Output status: {msg}")
                    else:
                        self.print_status("warning", f"Output status: {msg}")

                    return True
                else:
                    self.print_status("error", "Output file not created")
                    return False
            else:
                self.print_status("error", "Greetings generation failed")
                if result.stderr:
                    print(f"\nError details:\n{result.stderr}")
                return False

        except Exception as e:
            self.print_status("error", f"Failed to run Greetings Agent: {str(e)}")
            return False

    def run_agent_4_evaluation_sender(self, recipient_email: Optional[str] = None) -> bool:
        """
        Run Agent 4: Evaluation Gmail Sender - Send feedback to students.

        Args:
            recipient_email: Email to send to (for testing). If None, asks user.

        Returns:
            True if successful, False otherwise
        """
        self.print_header("Agent 4: Evaluation Gmail Sender - Send Feedback")

        # Check if input file exists
        if not self.check_file_exists(self.GREETINGS_OUTPUT):
            self.print_status("error", f"Input file not found: {self.GREETINGS_OUTPUT}")
            self.print_status("info", "Please run Agent 3 first")
            return False

        # Check if input is ready
        ready, msg = self.check_status_ready(self.GREETINGS_OUTPUT)
        if not ready:
            self.print_status("error", f"Input not ready: {msg}")
            return False

        self.print_status("success", f"Input ready: {msg}")

        # Ask for recipient email if not provided
        if not recipient_email:
            print(f"\nDefault recipient: {self.DEFAULT_EMAIL}")
            print("Enter a recipient email to send all drafts to that address (for testing)")
            print("Or press Enter to use the default email")
            print()
            recipient_email = input("Recipient email: ").strip()

            if not recipient_email:
                recipient_email = self.DEFAULT_EMAIL

        self.print_status("info", f"Recipient: {recipient_email}")

        # Confirm before sending
        print()
        confirm = input(f"Send email drafts to {recipient_email}? (yes/no): ").strip().lower()
        if confirm not in ["yes", "y"]:
            self.print_status("info", "Cancelled by user")
            return False

        try:
            # Run evaluation sender
            self.print_status("info", "Running Evaluation Gmail Sender...")

            cmd = [
                sys.executable,
                "-m", "evaluationAgent.cli",
                "create-drafts",
                "--input", str(self.GREETINGS_OUTPUT),
                "--to", recipient_email
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode == 0:
                self.print_status("success", "Email drafts created successfully")
                self.print_status("info", f"Drafts saved to: {self.EMAIL_DRAFTS_DIR}")
                return True
            else:
                self.print_status("error", "Email draft creation failed")
                if result.stderr:
                    print(f"\nError details:\n{result.stderr}")
                return False

        except Exception as e:
            self.print_status("error", f"Failed to run Evaluation Sender: {str(e)}")
            return False

    def run_all_agents(self) -> bool:
        """
        Run all agents in sequence (1 -> 2 -> 3 -> 4).

        Returns:
            True if all successful, False otherwise
        """
        self.print_header("Running All Agents Synchronously")

        # Gather inputs before starting
        print("\nBefore starting, we need to gather some information:\n")

        # Ask for Agent 1 filters
        print("=" * 70)
        print("STEP 1: Email Export Filters")
        print("=" * 70)
        print("\nSpecify how to filter emails from Gmail:")
        print("You can use any combination of filters below")
        print("\nExamples:")
        print("  Label:    'homework', 'lesson19', 'assignments'")
        print("  Subject:  'Homework 19', 'Assignment', 'Project'")
        print("  After:    '2025-12-01' (emails after this date)")
        print("  Before:   '2025-12-03' (emails before this date)")
        print("  Newer:    '7d' (last 7 days), '2m' (last 2 months)")
        print()

        label = input("Enter Gmail label (or press Enter to skip): ").strip()
        subject = input("Enter subject text (or press Enter to skip): ").strip()
        after = input("Enter after date YYYY-MM-DD (or press Enter to skip): ").strip()
        before = input("Enter before date YYYY-MM-DD (or press Enter to skip): ").strip()
        newer_than = input("Enter newer than (e.g., 7d, 2m) (or press Enter to skip): ").strip()

        if not any([label, subject, after, before, newer_than]):
            self.print_status("warning", "No filters specified. Will export all emails.")
            proceed = input("Continue? (yes/no): ").strip().lower()
            if proceed not in ["yes", "y"]:
                self.print_status("info", "Cancelled by user")
                return False

        # Ask for Agent 4 recipient email
        print("\n" + "=" * 70)
        print("STEP 2: Email Recipient Configuration")
        print("=" * 70)
        print(f"\nDefault recipient: {self.DEFAULT_EMAIL}")
        print("Enter a recipient email to send all drafts to that address (for testing)")
        print("Or press Enter to use the default email")
        print()
        recipient_email = input("Recipient email: ").strip()

        if not recipient_email:
            recipient_email = self.DEFAULT_EMAIL

        # Confirm configuration
        print("\n" + "=" * 70)
        print("CONFIGURATION SUMMARY")
        print("=" * 70)
        print(f"Gmail Label:    {label if label else '(none)'}")
        print(f"Subject Filter: {subject if subject else '(none)'}")
        print(f"After Date:     {after if after else '(none)'}")
        print(f"Before Date:    {before if before else '(none)'}")
        print(f"Newer Than:     {newer_than if newer_than else '(none)'}")
        print(f"Recipient:      {recipient_email}")
        print("=" * 70)
        print()

        confirm = input("Start processing with this configuration? (yes/no): ").strip().lower()
        if confirm not in ["yes", "y"]:
            self.print_status("info", "Cancelled by user")
            return False

        print()

        # Agent 1: GmailAgent
        if not self.run_agent_1_gmail(
            label=label,
            subject=subject,
            after=after,
            before=before,
            newer_than=newer_than
        ):
            self.print_status("error", "Agent 1 failed. Stopping workflow.")
            return False

        # Agent 2: Repository Analyzer
        if not self.run_agent_2_repo_analyzer():
            self.print_status("error", "Agent 2 failed. Stopping workflow.")
            return False

        # Agent 3: Personalized Greetings
        if not self.run_agent_3_greetings():
            self.print_status("error", "Agent 3 failed. Stopping workflow.")
            return False

        # Agent 4: Evaluation Sender
        if not self.run_agent_4_evaluation_sender(recipient_email=recipient_email):
            self.print_status("error", "Agent 4 failed. Stopping workflow.")
            return False

        self.print_header("All Agents Completed Successfully!")
        return True

    def reset_all(self):
        """Reset/clean all input and output files."""
        self.print_header("Reset - Cleaning All Inputs and Outputs")

        deleted_files = []
        deleted_dirs = []

        # Files to delete
        files_to_delete = [
            self.GMAIL_OUTPUT,
            self.REPO_ANALYZER_OUTPUT,
            self.GREETINGS_OUTPUT,
        ]

        for file_path in files_to_delete:
            if file_path.exists():
                file_path.unlink()
                deleted_files.append(file_path.name)
                self.print_status("success", f"Deleted: {file_path.name}")

        # Directories to clean (but keep the directory itself)
        dirs_to_clean = [
            self.EXPORTS_DIR,
            self.OUTPUT_DIR,
            self.GREETINGS_RESULTS_DIR,
            self.EMAIL_DRAFTS_DIR,
        ]

        for dir_path in dirs_to_clean:
            if dir_path.exists():
                # Delete all files in directory
                for item in dir_path.iterdir():
                    if item.is_file():
                        item.unlink()
                        deleted_files.append(f"{dir_path.name}/{item.name}")
                    elif item.is_dir():
                        shutil.rmtree(item)
                        deleted_dirs.append(f"{dir_path.name}/{item.name}")

        self.print_status("success", f"Deleted {len(deleted_files)} files")
        self.print_status("success", f"Deleted {len(deleted_dirs)} directories")
        self.print_status("info", "All outputs cleaned. Ready for fresh run.")

    def show_menu(self):
        """Display main menu."""
        print("\n" + "=" * 70)
        print("  EMAIL SKILL AGENTS ORCHESTRATOR")
        print("=" * 70)
        print("\n>> Automated Homework Grading Workflow\n")
        print("1. Run Email Generator (GmailAgent)")
        print("2. Run Repository Analyzer (on existing input)")
        print("3. Run Greetings Agent (on existing output)")
        print("4. Create Email Drafts and Send (Evaluation Sender)")
        print("5. Run ALL Agents (1 -> 2 -> 3 -> 4)")
        print("6. Reset - Clean All Inputs/Outputs")
        print("7. Exit")
        print("\n" + "-" * 70)

    def run(self):
        """Main orchestrator loop."""
        while True:
            self.show_menu()

            try:
                choice = input("\nSelect option (1-7): ").strip()

                if choice == "1":
                    self.run_agent_1_gmail()

                elif choice == "2":
                    self.run_agent_2_repo_analyzer()

                elif choice == "3":
                    self.run_agent_3_greetings()

                elif choice == "4":
                    self.run_agent_4_evaluation_sender()

                elif choice == "5":
                    self.run_all_agents()

                elif choice == "6":
                    confirm = input("\n[WARN] This will delete all outputs. Continue? (yes/no): ").strip().lower()
                    if confirm in ["yes", "y"]:
                        self.reset_all()
                    else:
                        self.print_status("info", "Reset cancelled")

                elif choice == "7":
                    self.print_status("info", "Exiting orchestrator. Goodbye!")
                    break

                else:
                    self.print_status("warning", "Invalid option. Please select 1-7.")

            except KeyboardInterrupt:
                print("\n")
                self.print_status("info", "Interrupted by user. Exiting...")
                break
            except Exception as e:
                self.print_status("error", f"Unexpected error: {str(e)}")

            # Pause before showing menu again
            input("\nPress Enter to continue...")


def main():
    """Entry point for orchestrator."""
    orchestrator = AgentOrchestrator()
    orchestrator.run()


if __name__ == "__main__":
    main()
