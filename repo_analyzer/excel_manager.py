"""
Excel manager for reading input and writing output files.
"""

import os
from pathlib import Path
from typing import List, Dict, Tuple
from datetime import datetime
import statistics

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")

from .config import EXCEL_COLUMNS, OUTPUT_DIR, STATUS_SUCCESS
from .errors import ExcelNotFoundError, ExcelInvalidFormatError


class ExcelManager:
    """Manage Excel input/output operations."""

    def read_input(self, file_path: str) -> List[Dict]:
        """
        Read Excel file and extract repository data.

        Args:
            file_path: Path to input Excel file

        Returns:
            List of dicts with keys: id, date, subject, url, row_index

        Raises:
            ExcelNotFoundError: If file doesn't exist
            ExcelInvalidFormatError: If file format is invalid
        """
        file_path = Path(file_path)

        # Check file exists
        if not file_path.exists():
            raise ExcelNotFoundError(f"Excel file not found: {file_path}")

        # Check extension
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ExcelInvalidFormatError(f"Not a valid Excel file: {file_path}")

        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Get headers from first row
            headers = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value:
                    headers[cell.value.strip()] = col_idx

            # Validate required columns
            required_cols = ['URL', 'Status']  # Minimum requirement
            missing_cols = [col for col in required_cols if col not in headers]

            if missing_cols:
                raise ExcelInvalidFormatError(
                    f"Missing required columns: {', '.join(missing_cols)}"
                )

            # Check if Status column exists and validate all rows are "ready"
            status_col = headers.get('Status')
            if status_col:
                all_ready = True
                not_ready_count = 0
                for row in ws.iter_rows(min_row=2, values_only=False):
                    status_value = row[status_col - 1].value
                    if status_value != 'ready':
                        all_ready = False
                        not_ready_count += 1

                if not all_ready:
                    raise ExcelInvalidFormatError(
                        f"Cannot process: {not_ready_count} row(s) have Status != 'ready'. "
                        f"Previous agent must complete successfully before this agent can run."
                    )

            # Extract data rows
            data = []
            url_col = headers.get('URL')
            id_col = headers.get('ID')
            date_col = headers.get('Date')
            subject_col = headers.get('Subject')

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                # Get URL (required)
                url = row[url_col - 1].value if url_col else None

                # Skip rows without URL
                if not url or not str(url).strip():
                    continue

                # Get optional fields
                row_data = {
                    'url': str(url).strip(),
                    'row_index': row_idx
                }

                if id_col:
                    row_data['id'] = row[id_col - 1].value or ''

                if date_col:
                    row_data['date'] = row[date_col - 1].value or ''

                if subject_col:
                    row_data['subject'] = row[subject_col - 1].value or ''

                data.append(row_data)

            if not data:
                raise ExcelInvalidFormatError("No valid repository URLs found in Excel")

            return data

        except Exception as e:
            if isinstance(e, (ExcelNotFoundError, ExcelInvalidFormatError)):
                raise
            raise ExcelInvalidFormatError(f"Failed to read Excel file: {str(e)}")

    def write_output(
        self,
        output_path: str,
        original_data: List[Dict],
        results: List[Dict],
        processing_time: float
    ):
        """
        Write graded results to Excel file.

        Args:
            output_path: Path to output Excel file
            original_data: Original data from input Excel
            results: Processing results for each repository
            processing_time: Total processing time in seconds

        Raises:
            ExcelInvalidFormatError: If writing fails
        """
        try:
            # Create workbook
            wb = Workbook()

            # Create results sheet
            self._create_results_sheet(wb, original_data, results)

            # Create summary sheet
            self._create_summary_sheet(wb, results, processing_time)

            # Ensure output directory exists
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            wb.save(output_path)

        except Exception as e:
            raise ExcelInvalidFormatError(f"Failed to write Excel file: {str(e)}")

    def _create_results_sheet(
        self,
        wb: Workbook,
        original_data: List[Dict],
        results: List[Dict]
    ):
        """
        Create the graded results sheet.

        Args:
            wb: Workbook object
            original_data: Original input data
            results: Processing results
        """
        # Get or create sheet
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = 'Graded Results'
        else:
            ws = wb.create_sheet('Graded Results')

        # Define columns
        columns = EXCEL_COLUMNS['output']
        headers = [
            columns['id'],
            columns['date'],
            columns['subject'],
            columns['url'],
            columns['grade'],
            columns['total_files'],
            columns['files_under_130'],
            columns['total_lines'],
            columns['status'],
            columns['error'],
            'Status'  # Added Status column for agent coordination
        ]

        # Write headers
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write data rows
        for orig, result in zip(original_data, results):
            row = [
                orig.get('id', ''),
                orig.get('date', ''),
                orig.get('subject', ''),
                orig.get('url', ''),
                result.get('grade', 0.0),
                result.get('total_files', 0),
                result.get('files_under_130', 0),
                result.get('total_lines', 0),
                result.get('status', ''),
                result.get('error', ''),
                'ready'  # Set Status to "ready" after processing
            ]
            ws.append(row)

        # Format columns
        ws.column_dimensions['A'].width = 10  # ID
        ws.column_dimensions['B'].width = 20  # Date
        ws.column_dimensions['C'].width = 30  # Subject
        ws.column_dimensions['D'].width = 50  # URL
        ws.column_dimensions['E'].width = 12  # Grade
        ws.column_dimensions['F'].width = 12  # Total Files
        ws.column_dimensions['G'].width = 12  # Files <130
        ws.column_dimensions['H'].width = 12  # Total Lines
        ws.column_dimensions['I'].width = 15  # Status
        ws.column_dimensions['J'].width = 40  # Error
        ws.column_dimensions['K'].width = 15  # Status (agent coordination)

        # Color-code status
        for row_idx in range(2, len(original_data) + 2):
            status_cell = ws[f'I{row_idx}']

            if status_cell.value == STATUS_SUCCESS:
                status_cell.fill = PatternFill(
                    start_color='90EE90',
                    end_color='90EE90',
                    fill_type='solid'
                )
            else:
                status_cell.fill = PatternFill(
                    start_color='FFB6C1',
                    end_color='FFB6C1',
                    fill_type='solid'
                )

            # Color-code agent coordination Status column (K)
            agent_status_cell = ws[f'K{row_idx}']
            if agent_status_cell.value == 'ready':
                agent_status_cell.fill = PatternFill(
                    start_color='90EE90',
                    end_color='90EE90',
                    fill_type='solid'
                )

    def _create_summary_sheet(
        self,
        wb: Workbook,
        results: List[Dict],
        processing_time: float
    ):
        """
        Create the summary statistics sheet.

        Args:
            wb: Workbook object
            results: Processing results
            processing_time: Total processing time in seconds
        """
        ws = wb.create_sheet('Summary')

        # Calculate statistics
        total_repos = len(results)
        successful = sum(1 for r in results if r.get('status') == STATUS_SUCCESS)
        failed = total_repos - successful

        # Calculate grade statistics (only for successful repos)
        grades = [r.get('grade', 0.0) for r in results if r.get('status') == STATUS_SUCCESS]

        if grades:
            avg_grade = statistics.mean(grades)
            median_grade = statistics.median(grades)
            min_grade = min(grades)
            max_grade = max(grades)
            std_dev = statistics.stdev(grades) if len(grades) > 1 else 0.0
        else:
            avg_grade = median_grade = min_grade = max_grade = std_dev = 0.0

        # Calculate grade distribution
        distribution = {
            '0-20%': sum(1 for g in grades if 0 <= g < 20),
            '20-40%': sum(1 for g in grades if 20 <= g < 40),
            '40-60%': sum(1 for g in grades if 40 <= g < 60),
            '60-80%': sum(1 for g in grades if 60 <= g < 80),
            '80-100%': sum(1 for g in grades if 80 <= g <= 100)
        }

        # Write summary
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws.append([f"Grading Summary - Generated {timestamp}"])
        ws.append([])
        ws.append(["Total Repositories:", total_repos])
        ws.append(["Successfully Graded:", successful])
        ws.append(["Failed:", failed])
        ws.append([])
        ws.append(["Grade Statistics:"])
        ws.append(["  Average:", f"{avg_grade:.2f}%"])
        ws.append(["  Median:", f"{median_grade:.2f}%"])
        ws.append(["  Min:", f"{min_grade:.2f}%"])
        ws.append(["  Max:", f"{max_grade:.2f}%"])
        ws.append(["  Std Dev:", f"{std_dev:.2f}%"])
        ws.append([])
        ws.append(["Grade Distribution:"])

        for range_label, count in distribution.items():
            percentage = (count / len(grades) * 100) if grades else 0
            ws.append([f"  {range_label}:", f"{count} repos ({percentage:.1f}%)"])

        ws.append([])
        ws.append(["Processing Time:"])
        minutes = int(processing_time // 60)
        seconds = int(processing_time % 60)
        ws.append([f"  {minutes} minutes {seconds} seconds"])

        # Error report
        if failed > 0:
            ws.append([])
            ws.append(["Error Report:"])

            for result in results:
                if result.get('status') != STATUS_SUCCESS:
                    error_msg = result.get('error', 'Unknown error')
                    url = result.get('url', 'Unknown URL')
                    ws.append([f"  - {url}: {error_msg}"])

        # Style title
        ws['A1'].font = Font(bold=True, size=14)

        # Auto-width columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    def generate_output_path(self, input_path: str) -> str:
        """
        Generate output file path based on input path.

        Args:
            input_path: Path to input Excel file

        Returns:
            Path to output Excel file (in output/ directory)
        """
        input_path = Path(input_path)
        stem = input_path.stem
        output_filename = f"{stem}_graded.xlsx"

        # Create output directory if it doesn't exist
        output_dir = Path(OUTPUT_DIR)
        output_dir.mkdir(exist_ok=True)

        return str(output_dir / output_filename)
