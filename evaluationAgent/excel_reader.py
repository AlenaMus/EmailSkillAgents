"""
Excel file reader for Evaluation Gmail Sender Agent.

This module handles reading and parsing the Excel file with student grades
and personalized greetings from the previous agent.
"""

from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .config import EXCEL_COLUMNS
from .errors import (
    ExcelFileNotFoundError,
    ExcelReadError,
    ExcelValidationError,
)


class StudentData:
    """Data class representing a student's grading information."""

    def __init__(
        self,
        id: str,
        date: str,
        subject: str,
        url: str,
        grade: float,
        total_files: int,
        files_under_130: int,
        total_lines: int,
        status: str,
        error_message: Optional[str],
        personalized_greeting: str,
        greeting_persona: str,
    ):
        self.id = id
        self.date = date
        self.subject = subject
        self.url = url
        self.grade = float(grade) if grade is not None else 0.0
        self.total_files = int(total_files) if total_files is not None else 0
        self.files_under_130 = int(files_under_130) if files_under_130 is not None else 0
        self.total_lines = int(total_lines) if total_lines is not None else 0
        self.status = status
        self.error_message = error_message
        self.personalized_greeting = personalized_greeting
        self.greeting_persona = greeting_persona

    def is_valid_for_email(self) -> bool:
        """
        Check if this student data is valid for sending email.

        Returns:
            True if student has valid grade and successful status
        """
        return (
            self.status == 'Success' and
            self.grade > 0 and
            self.personalized_greeting and
            len(self.personalized_greeting.strip()) > 0 and
            self.url and
            self.url.startswith(('http://', 'https://'))
        )

    def get_skip_reason(self) -> Optional[str]:
        """
        Get reason why this student should be skipped.

        Returns:
            Reason string if should be skipped, None if valid
        """
        if self.status != 'Success':
            return f"Status: {self.status}"

        if self.grade <= 0:
            return "No valid grade (0%)"

        if not self.personalized_greeting or len(self.personalized_greeting.strip()) == 0:
            return "No personalized greeting"

        if not self.url or not self.url.startswith(('http://', 'https://')):
            return "Invalid repository URL"

        return None

    def __repr__(self) -> str:
        return f"StudentData(id={self.id[:8]}..., grade={self.grade}%, status={self.status})"


class ExcelReader:
    """Reader for Excel files containing student grading data."""

    def __init__(self, filepath: str):
        """
        Initialize Excel reader.

        Args:
            filepath: Path to Excel file

        Raises:
            ExcelFileNotFoundError: If file doesn't exist
        """
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise ExcelFileNotFoundError(str(self.filepath))

    def read(self) -> List[StudentData]:
        """
        Read Excel file and parse student data.

        Returns:
            List of StudentData objects

        Raises:
            ExcelReadError: If file cannot be read
            ExcelValidationError: If file structure is invalid
        """
        try:
            workbook = openpyxl.load_workbook(self.filepath)
            worksheet = workbook.active

            # Validate headers
            self._validate_headers(worksheet)

            # Parse rows
            students = self._parse_rows(worksheet)

            workbook.close()

            return students

        except ExcelValidationError:
            raise
        except Exception as e:
            raise ExcelReadError(str(self.filepath), str(e))

    def _validate_headers(self, worksheet: Worksheet) -> None:
        """
        Validate that Excel file has required columns.

        Args:
            worksheet: Excel worksheet

        Raises:
            ExcelValidationError: If required columns are missing
        """
        # Get headers from first row
        headers = [cell.value for cell in worksheet[1]]

        # Check for required columns
        required_columns = [
            EXCEL_COLUMNS['id'],
            EXCEL_COLUMNS['subject'],
            EXCEL_COLUMNS['url'],
            EXCEL_COLUMNS['grade'],
            EXCEL_COLUMNS['status'],
            EXCEL_COLUMNS['personalized_greeting'],
        ]

        missing_columns = [col for col in required_columns if col not in headers]

        if missing_columns:
            raise ExcelValidationError(
                f"Missing required columns: {', '.join(missing_columns)}",
                missing_columns=missing_columns
            )

        # Check if agent coordination Status column exists and validate all rows are "ready"
        if 'Status' in headers:
            # Find the rightmost Status column (agent coordination Status)
            status_indices = [i for i, h in enumerate(headers) if h == 'Status']
            if status_indices:
                status_col_idx = status_indices[-1] + 1  # Get rightmost Status column
                all_ready = True
                not_ready_count = 0
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if status_col_idx <= len(row):
                        status_value = row[status_col_idx - 1]
                        if status_value != 'ready':
                            all_ready = False
                            not_ready_count += 1

                if not all_ready:
                    raise ExcelValidationError(
                        f"Cannot process: {not_ready_count} row(s) have Status != 'ready'. "
                        f"Previous agent (Personalized Greetings Agent) must complete successfully before this agent can run."
                    )

    def _parse_rows(self, worksheet: Worksheet) -> List[StudentData]:
        """
        Parse rows from worksheet into StudentData objects.

        Args:
            worksheet: Excel worksheet

        Returns:
            List of StudentData objects
        """
        # Get headers
        headers = [cell.value for cell in worksheet[1]]

        # Create column index map
        # For duplicate headers (like Status), use the FIRST occurrence
        # (Repository Analyzer Status, not agent coordination Status)
        col_map = {}
        for idx, header in enumerate(headers):
            if header not in col_map:  # Only store first occurrence
                col_map[header] = idx

        students = []

        # Parse data rows (skip header row)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            try:
                student = StudentData(
                    id=self._get_cell_value(row, col_map, 'id', ''),
                    date=self._get_cell_value(row, col_map, 'date', ''),
                    subject=self._get_cell_value(row, col_map, 'subject', ''),
                    url=self._get_cell_value(row, col_map, 'url', ''),
                    grade=self._get_cell_value(row, col_map, 'grade', 0),
                    total_files=self._get_cell_value(row, col_map, 'total_files', 0),
                    files_under_130=self._get_cell_value(row, col_map, 'files_under_130', 0),
                    total_lines=self._get_cell_value(row, col_map, 'total_lines', 0),
                    status=self._get_cell_value(row, col_map, 'status', ''),
                    error_message=self._get_cell_value(row, col_map, 'error_message', None),
                    personalized_greeting=self._get_cell_value(row, col_map, 'personalized_greeting', ''),
                    greeting_persona=self._get_cell_value(row, col_map, 'greeting_persona', ''),
                )
                students.append(student)
            except Exception as e:
                # Log warning but continue processing
                print(f"Warning: Failed to parse row: {e}")
                continue

        return students

    def _get_cell_value(
        self,
        row: tuple,
        col_map: Dict[str, int],
        key: str,
        default: any = None
    ) -> any:
        """
        Get cell value from row by column key.

        Args:
            row: Row tuple
            col_map: Column name to index mapping
            key: Column key from EXCEL_COLUMNS
            default: Default value if not found

        Returns:
            Cell value or default
        """
        col_name = EXCEL_COLUMNS.get(key)
        if not col_name or col_name not in col_map:
            return default

        col_idx = col_map[col_name]
        if col_idx >= len(row):
            return default

        value = row[col_idx]
        return value if value is not None else default


def read_students_from_excel(filepath: str) -> List[StudentData]:
    """
    Convenience function to read students from Excel file.

    Args:
        filepath: Path to Excel file

    Returns:
        List of StudentData objects

    Raises:
        ExcelFileNotFoundError: If file doesn't exist
        ExcelReadError: If file cannot be read
        ExcelValidationError: If file structure is invalid
    """
    reader = ExcelReader(filepath)
    return reader.read()


def filter_valid_students(students: List[StudentData]) -> tuple[List[StudentData], List[StudentData]]:
    """
    Filter students into valid and invalid lists.

    Args:
        students: List of all students

    Returns:
        Tuple of (valid_students, invalid_students)
    """
    valid = [s for s in students if s.is_valid_for_email()]
    invalid = [s for s in students if not s.is_valid_for_email()]
    return valid, invalid
