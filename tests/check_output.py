from openpyxl import load_workbook

wb = load_workbook('output/homework_emails_graded.xlsx')

# Check Results sheet
print('=== GRADED RESULTS ===')
ws = wb['Graded Results']
print('Headers:')
print([cell.value for cell in ws[1]])
print('\nData rows:')
for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
    print(row)

# Check Summary sheet
print('\n=== SUMMARY ===')
ws_summary = wb['Summary']
for row in ws_summary.iter_rows(values_only=True):
    if row[0]:
        print(row)
